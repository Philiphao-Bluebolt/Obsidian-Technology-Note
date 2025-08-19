from datetime import datetime
import calendar
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, Color, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class TaskTable:
    def __init__(self):
        self.wb = load_workbook("./任务时间表.xlsx")

        # Weekday cells style
        self.begin_weekday_row = "B1"
        self.begin_date_row = "B2"

        self.weekday_bg_colors = {
            1: "#F9C3BD",
            2: "#F2E2A8",
            3: "#EFF8C0",
            4: "#BBF3C6",
            5: "#B7DCF7",
            6: "#C8CFFC",
            7: "#E5C1E7"
        }

        self.weekday_txt_colors = {
            1: "#000000", 
            2: "#000000",
            3: "#000000",
            4: "#000000",
            5: "#000000",
            6: "#F0ECE0",
            7: "#F0ECE0"
        }

        self.weekday_to_num = {
            "Mon": 1,
            "Tue": 2,
            "Wed": 3,
            "Thu": 4,
            "Fri": 5,
            "Sat": 6,
            "Sun": 7
        }

        # Hour cells style
        self.rest_hours = [6, 12, 13, 18, 19, 23]
        self.begin_hour = 6
        self.hour_cells = "A3:A20"

        self.rest_hour_style = NamedStyle(name="rest_hour")
        self.rest_hour_style.font = Font(name="等线", size="11", color="757171")
        self.rest_hour_style.fill = PatternFill(start_color="F2F2F2")
        self.rest_hour_style.number_format = "h:mm"

        self.work_hour_style = NamedStyle(name="work_hour")
        self.work_hour_style.font = Font(name="等线", size="11", color="302C46")
        self.work_hour_style.fill = PatternFill(start_color="E7E6E6")
        self.work_hour_style.number_format = "h:mm"

    def add_sheet(self, year=datetime.now().year, month=datetime.now().month): 

        year_str = year - 2000
        month_str = month if month >= 10 else "0" + str(month)
        ws = self.wb.create_sheet(f"{year}-{month_str}", 0)
    
        # Set the width of columns and height of rows
        for row in range(1, 3): ws.row_dimensions[row].height = 20
        for row in range(3, 21): ws.row_dimensions[row].height = 19.1 
        for col in range(2, 33): ws.column_dimensions[get_column_letter(col)].width = 22.1      
        ws.column_dimensions['A'].width = 15  

        # Set the style of hour cells
        hour_cnt = self.begin_hour
        for rows in ws[self.hour_cells]:
            for cell in rows:
                cell.style = self.rest_hour_style if hour_cnt in self.rest_hours else self.work_hour_style
                cell.value = hour_cnt
                hour_cnt += 1
                
        # Set the style of day cells
        first_weekday, num_days = calendar.monthrange(year, month)
        first_weekday = calendar.day_abbr[first_weekday]

        weekday_cnt = self.weekday_to_num[first_weekday]
        weekday_cell = ws[self.begin_weekday_row]
        date_cell = ws[self.begin_date_row]
        # for _ in range(num_days):
            
        #     weekday_cell =
            
            

            

        # Set the style of record cells



        # Saving the changes
        self.wb.save("./任务时间表.xlsx")
        return ws

    def add_sheet_next_month(self):
        # Create new sheet only at the creating day
        now = datetime.now()
        year = now.year
        month = now.month
        create_day = 15
        if now.day != create_day:
            raise ValueError(f"Not on the creating day: {create_day}th")
        else: 
        # Move to next month
            if month == 12:
                month = 1
                year += 1
            else:
                month += 1
        self.add_sheet(month, year)

if __name__ == "__main__":
    
    table = TaskTable()
    table.add_sheet_next_month()
    


